import csv
import json
from urllib import parse
from bs4 import BeautifulSoup
import os
import urllib.request
from openpyxl import load_workbook, Workbook
import datetime
import inspect

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for data in _DATA:
            sheet.append(data)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        '''
        #컬럼 열너비 지정(지정 안하면 기본사이즈)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 50
        '''
        book.save(_FILENAME)

# 파일 다운로드 함수 -------------------------------------------
def file_down(_url, _savename):
    urllib.request.urlretrieve(_url, _savename)

# 날짜시간 스탬프 함수 ---------------------------------------------
def timeNow():
    t = ['(월)', '(화)', '(수)', '(목)', '(금)', '(토)', '(일)']
    r = datetime.datetime.today().weekday()
    return (datetime.datetime.now().strftime('%Y%m%d') + t[r])

def timeNow_Second():
    t = ['(월)', '(화)', '(수)', '(목)', '(금)', '(토)', '(일)']
    r = datetime.datetime.today().weekday()
    return (datetime.datetime.now().strftime('%Y%m%d') + t[r] + datetime.datetime.now().strftime('_%H%M%S'))

#   엑셀내용을 채우는 함수
def excel_description(home,project,index,title,addr,kategorie1,kategorie2,attribute1,attribute2,img_url,url_origin,img_url_origin,else_text):
    #엑셀을 채우기위한 리스트 생성
    dataList = []
    baseUrl = home
    DIR = project  # 이미지저장할 폴더

    # 엑셀파일 및 작업기록 파일생성 -------------------------
    FILENAME = DIR + "_crawling_" + timeNow() + ".xlsx"  # 생성될 엑셀파일명
    HEADER = ['순번',
              '업체명',
              '제목',
              '인테리어장소',
              '카테고리1',
              '카테고리2',
              '속성_공간',
              '속성_스타일',
              '속성_평수',
              '해시태그',
              '이미지저장주소',
              '추출페이지주소',
              '원본이미지주소',
              '기타_텍스트']

    #   엑셀파일을 미리 만들어줘야 첫번째 데이터를 안넘기고 넣을수있다
    save_excel(FILENAME, None, HEADER)  # 파일명과 헤더 삽입하여 엑셀파일 생성

    #   함수 파라미터값으로 데이터를 채워준다
    data_index = index #순번
    data_pro = project  # 업체명,\
    data_title = title  # 제목
    data_addr = addr  # 인테리어장소
    data_category01 = kategorie1  # 카테고리1
    data_category02 = kategorie2  # 카테고리2
    data_filter01 = attribute1  # 속성_공간
    data_filter02 = attribute2  # 속성_스타일
    data_filter03 = ''  # 속성_미정
    data_hashTag = ''  # 해시태그
    data_img_url = img_url  # 이미지저장주소
    data_url_origin = url_origin  # 추출페이지주소
    data_img_url_origin = img_url_origin  # 원본이미지주소
    data_etcText = else_text  # 기타_텍스트

    #   엑셀을채위기위한 리스트에 추가
    dataList.append([
        data_index,
        data_pro,
        data_title,
        data_addr,
        data_category01,
        data_category02,
        data_filter01,
        data_filter02,
        data_filter03,
        data_hashTag,
        data_img_url,
        data_url_origin,
        data_img_url_origin,
        data_etcText
    ])

    #   엑셀에 데이터추가
    save_excel(FILENAME, dataList, HEADER)  # 파일명과 헤더 삽입하여 엑셀파일 생성

def create_excel(folder_idx,project):
    DIR = folder_idx + "." + project + "_사진자료/"  # 이미지저장할 폴더
    # 엑셀파일 및 작업기록 파일생성 -------------------------
    FILENAME =  DIR + project + "_crawling_" + timeNow() + ".xlsx"  # 생성될 엑셀파일명
    HEADER = ['순번',
              '업체명',
              '제목',
              '인테리어장소',
              '카테고리1',
              '카테고리2',
              '속성_공간',
              '속성_스타일',
              '속성_평수',
              '해시태그',
              '이미지저장주소',
              '추출페이지주소',
              '원본이미지주소',
              '기타_텍스트']

    #   엑셀파일을 미리 만들어줘야 첫번째 데이터를 안넘기고 넣을수있다
    save_excel(FILENAME, None, HEADER)  # 파일명과 헤더 삽입하여 엑셀파일 생성

def data_save(folder_idx,project,dataList):
    DIR = folder_idx + "." + project + "_사진자료/"  # 이미지저장할 폴더
    FILENAME =  DIR + project + "_crawling_" + timeNow() + ".xlsx"  # 생성될 엑셀파일명
    save_excel(FILENAME, dataList, None)  # 파일명과 헤더 삽입하여 엑셀파일 생성



def excel_description_resources(home,project,index,category1,category2,category3,title,price,img_url):
    #엑셀을 채우기위한 리스트 생성
    dataList = []
    baseUrl = home
    DIR = project  # 이미지저장할 폴더

    # 엑셀파일 및 작업기록 파일생성 -------------------------
    FILENAME = DIR + "_crawling_" + ".xlsx"  # 생성될 엑셀파일명
    HEADER = ['순번',
              '카테고리1',
              '카테고리2',
              '카테고리3',
              '제품이름',
              '제품단가',
              '제품사진'
              ]

    #   엑셀파일을 미리 만들어줘야 첫번째 데이터를 안넘기고 넣을수있다
    save_excel(FILENAME, None, HEADER)  # 파일명과 헤더 삽입하여 엑셀파일 생성

    #   함수 파라미터값으로 데이터를 채워준다
    data_index = index #순번
    data_pro = project  # 업체명,\
    data_category01 = category1  # 카테고리1
    data_category02 = category2  # 카테고리2
    data_category03 = category3  # 카테고리2
    data_title = title  # 제목
    data_price = price
    data_img_url = img_url
    #   엑셀을채위기위한 리스트에 추가
    dataList.append([
        data_index,
        data_pro,
        data_category01,
        data_category02,
        data_category03,
        data_title,
        data_price,
        data_img_url
    ])

    #   엑셀에 데이터추가
    save_excel(FILENAME, dataList, HEADER)  # 파일명과 헤더 삽입하여 엑셀파일 생성

def create_excel_resources(folder_idx,project):
    DIR = folder_idx + "." + project + "_사진자료/"  # 이미지저장할 폴더
    # 엑셀파일 및 작업기록 파일생성 -------------------------
    FILENAME =  DIR + project + "_crawling_" + ".xlsx"  # 생성될 엑셀파일명
    HEADER = ['순번',
              '카테고리1',
              '카테고리2',
              '카테고리3',
              '제품이름',
              '제품단가',
              '제품사진'
              ]

    #   엑셀파일을 미리 만들어줘야 첫번째 데이터를 안넘기고 넣을수있다
    save_excel(FILENAME, None, HEADER)  # 파일명과 헤더 삽입하여 엑셀파일 생성

def data_save_resources(folder_idx,project,dataList):
    DIR = folder_idx + "." + project + "_사진자료/"  # 이미지저장할 폴더
    FILENAME =  DIR + project + "_crawling_" + ".xlsx"  # 생성될 엑셀파일명
    save_excel(FILENAME, dataList, None)  # 파일명과 헤더 삽입하여 엑셀파일 생성